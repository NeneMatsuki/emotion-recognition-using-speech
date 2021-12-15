from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('prediction.xlsx')

wb.remove(wb['predictions'])
wb.create_sheet('predictions')
ws = wb['predictions']

for i in range(1,4):
    for j in range(1,4):
        column_letter = get_column_letter(j)
        ws[column_letter + str(i)] = "bye"

wb.save('prediction.xlsx')

