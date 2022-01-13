from charset_normalizer import models
from emotion_recognition import EmotionRecognizer
from deep_emotion_recognition import DeepEmotionRecognizer
import json
import os
from sys import byteorder
import sys
from struct import pack
from sklearn.ensemble import GradientBoostingClassifier, BaggingClassifier, RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
import time
import librosa
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from multiple_functions import *

from utils import get_best_estimators

def get_estimators_name(estimators):
    result = [ '"{}"'.format(estimator.__class__.__name__) for estimator, _, _ in estimators ]
    return ','.join(result), {estimator_name.strip('"'): estimator for estimator_name, (estimator, _, _) in zip(result, estimators)}

if __name__ == "__main__":
    estimators = get_best_estimators(True)
    estimators_str, estimator_dict = get_estimators_name(estimators)
    print(estimators_str)

    with open('predict.json', 'r') as config_file:
        data = json.load(config_file)
        mandatory_settings =    data["Mandatory Settings"][0]
        Test_or_train_mode =    mandatory_settings["Test or Train"].lower()

        # check if train or test, if not then exit
        if((Test_or_train_mode!= "train") and (Test_or_train_mode != "test")) :
            sys.exit("Please choose whether to Test or Train.\n This can be done under Mandatory Settings in predict.json")       
    
        # load mandatory settings
        model =     mandatory_settings["model"].format(estimators_str)
        frequency_features = mandatory_settings["frequency_features"]
        emotions =  mandatory_settings['emotions'].split(",")
        features =  mandatory_settings["features"].split(",")
        model_name = os.path.join(frequency_features,model)

        # if testing
        if(Test_or_train_mode == "test"):

            # load testing settings
            test_settings = data["Test settings"][0]
            test_mode =     test_settings["Test mode"].lower()

            # create detector instance
            if(model == "DecisionTreeClassifier"):
                detector = EmotionRecognizer(model = DecisionTreeClassifier() , emotions=emotions, model_name = model_name, features=features, verbose=0)
            else:
                detector = EmotionRecognizer(estimator_dict[model] , emotions=emotions, model_name = model_name, features=features, verbose=0)

            # if predicting a single audio
            if(test_mode == 'single'):

                # load settings and record length odf audio
                single_settings = test_settings["Test single"][0]
                audio = single_settings["Audio directory"]
                print(f'\nChosen to test a single audio using {model} trained on {frequency_features}')
                print(f'Length of audio: {librosa.get_duration(filename = audio)} seconds')

                #predict from filename passed in args
                start_predict = time.perf_counter()
                result = detector.predict_proba(audio)
                end_predict = time.perf_counter()

                # print result
                print(f'\n{result}')
                maximum = max(result, key=result.get)
                max_value = result[maximum]
                del result[maximum]
                second = max(result, key=result.get)
                second_value = result[second]

                print(f"\nfirst prediction  : {maximum} \nsecond prediction : {second} \ndifference is {(max_value - second_value)*100} %")
                print(f"\nTime it took to predict: {(end_predict - start_predict)*1000} ms")

            # if predicting multiple audio
            elif(test_mode == 'multiple'):
                multiple_settings = test_settings["Test multiple"][0]
                output = multiple_settings["output"]

                if(output == "excel"):
                    # initialise workbook 
                    wb = load_workbook('predict_from_audio/prediction.xlsx')
                    wb.remove(wb['predictions'])
                    wb.create_sheet('predictions')
                    sheet = wb['predictions']
                    sheet["A1"] = "True emotion"
                    sheet["B1"] = "Intensity" 
                    sheet["C1"] = "Result"

                    for i in range(len(emotions)):
                        sheet[get_column_letter(i + 4) + "1"] =  emotions[i]

                    rows = sm_predict_all_excel(detector = detector,rows = 2, cols = 1, sheet = sheet)
                    rows = predict_all_excel(detector = detector,rows = 2, cols = 1, sheet = sheet, file = 'test_tess_ravdess.csv')
                    rows = predict_all_excel(detector = detector,rows = 2, cols = 1, sheet = sheet, file = 'test_emodb.csv')                    
                    # rows = predict_excel(frequency = frequency_features[:3], detector = detector, folder = "Nene", rows = rows, cols = 1, sheet = sheet)
                    # rows = predict_excel(frequency = frequency_features[:3], detector = detector, folder = "JL", rows = rows, cols = 1, sheet = sheet)

                    wb.save('predict_from_audio/prediction.xlsx')
                    print('predictions saved to predict_from_audio/prediction.xlsx')
        
                # else if outputting to text
                else:
                    with open(file = 'predict_from_audio' + os.sep + 'predictions.txt', mode  = 'w') as file:
                        sm_predict_text(frequency= frequency_features[:3], detector = detector, emotions = emotions, file = file)
                        print('predictions saved to predict_from_audio/prediction.txt')
            
            # if single or multiple is not chosen
            else:
                sys.exit("Please choose whether to predict single or multiple.\n This can be done under Testing Settings, Test mode in predict.json")
        
        else:
            test_settings = data["Test settings"][0]
            train_classifiers = test_settings['Classifiers to train']
            start_train = time.perf_counter()
            
            for model in train_classifiers:
                    for model in models:

                        if(model == "DecisionTreeClassifier"):
                            detector = EmotionRecognizer(model = DecisionTreeClassifier() , emotions=emotions, model_name = model_name, features=features, verbose=0)
                        
                        else:
                            detector = EmotionRecognizer(estimator_dict[model] , emotions=emotions, model_name = model_name, features=features, verbose=0)
                        
                        # train the model and display status
                        detector.train()
                        print(f"\n{model} trained")
                        print(detector.confusion_matrix())
                        print("Test accuracy score: {:.3f}%".format(detector.test_score()*100))

            end_train = time.perf_counter()
            print(f"\nThis process took {end_train - start_train} seconds")
    


                

    