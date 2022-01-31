from emotion_recognition import EmotionRecognizer
import os
import sys
import time
import librosa
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from emotion_recognizer_functions import *

from utils import get_pre_tuned_models, get_pre_tuned_models_dict,load_input_settings
from live_voice_emotion_recognizer import record_without_file


if __name__ == "__main__":

    # load and check if input json file exists
    json_file = "test_train_input_config.json"
    if(not os.path.isfile(json_file)):
        print(f"config file, {json_file} does not exist")
        sys.exit(f"config file, {json_file} does not exist")

    # load mandatory settings
    test_or_train, mode, model_name, model_folder, emotions, features, model_dir, mode_settings = load_input_settings(json_file)

    # If test
    if(test_or_train == "test"):

        # create detector instance
        detector = EmotionRecognizer(emotions=emotions, model_dir = model_dir, features=features, verbose=0)

        # if predicting a single audio
        if(mode == 'single'):

            # get audio directory
            audio_dir = mode_settings["Audio directory"]

            # print to user what the settings are 
            print(f'\nChosen to test a single audio using {model_name} trained on {model_folder}')
            print(f'Length of audio: {librosa.get_duration(filename = audio_dir)} second_likely_emotion_keys')

            #predict from filename passed in args
            start_predict = time.perf_counter()
            emotion_probabilities = detector.predict_proba_file(audio_dir)
            end_predict = time.perf_counter()

            # print emotion_probabilitiess 
            print(f'\n{emotion_probabilities}')

            # get most likely emotion
            most_likely_emotion_key     = max(emotion_probabilities, key=emotion_probabilities.get) 
            most_likely_emotion_value   = emotion_probabilities[most_likely_emotion_key]
            del emotion_probabilities[most_likely_emotion_key]
        
            # get second most likely emotion
            second_likely_emotion_key   = max(emotion_probabilities, key=emotion_probabilities.get)
            second_likely_emotion_value = emotion_probabilities[second_likely_emotion_key]

            effective_prob = (most_likely_emotion_value - second_likely_emotion_value)*100
            # print to user 
            print(f"\nBest prediction  : {most_likely_emotion_key} \nsecond best prediction : {second_likely_emotion_key} \ndifference is {effective_prob} %")
            print(f"\nTherefore predicted {most_likely_emotion_key} with an effective probability of {effective_prob} %")
            print(f"Time it took to predict: {(end_predict - start_predict)*1000} ms")

        # if predicting multiple audio
        elif(mode == 'multiple'):

            # get the method to display the predictions  
            prediction_output_mode = mode_settings["Display predictions"]
            testing_portion = mode_settings["Testing portion"]

            #get boolean to plot stats
            if(mode_settings["Plot stats bool"].lower() == "true"):
                is_plot_stats = True
            elif(mode_settings["Plot stats bool"].lower() == "false"):
                is_plot_stats = False
            else:
                print("Please choose true or false in Plot stats bool, TEST MULTIPLE SETING")
                sys.exit("Please choose true or false in Plot stats bool, TEST MULTIPLE SETING")

            print(f'\nChosen to test multiple audio files using {model_name} trained on {model_folder}')

            # if display is excel and predicting al4l 10521 test dataset
            if(prediction_output_mode == "excel"):
                # initialise workbook 
                wb = load_workbook('test_audio/prediction.xlsx') # audio path to excel file
                wb.remove(wb['predictions'])
                wb.create_sheet('predictions')
                sheet = wb['predictions']
                sheet["A1"] = "True emotion"
                sheet["B1"] = "Intensity" 
                sheet["C1"] = "emotion_probabilities"

                # add labels to the workbook
                for i in range(len(emotions)):
                    sheet[get_column_letter(i + 4) + "1"] =  emotions[i]
                
                sheet[get_column_letter(i + 4) + "1"] = "audio file"

                # if testing all
                if(testing_portion == "all"):

                    # read through all files, and record the predicitons
                    rows, time_taken, duration = predict_all_excel(detector = detector,rows = 2, cols = 1, sheet = sheet, file = 'audio_csv/test_custom.csv', time_taken = [], duration = [])
                    rows, time_taken, duration = predict_all_excel(detector = detector,rows = rows, cols = 1, sheet = sheet, file = 'audio_csv/test_tess_ravdess.csv', time_taken = time_taken, duration = duration)
                    rows, time_taken, duration = predict_all_excel(detector = detector,rows = rows, cols = 1, sheet = sheet, file = 'audio_csv/test_emodb.csv', time_taken = time_taken, duration = duration)                    

                    # save and display where the predictions ar saved
                    wb.save('test_audio/prediction.xlsx')
                    print('predictions saved to test_audio/prediction.xlsx')

                    # if plotting the time taken is selected, plot the time taken to predict with the length of the audio
                    if (is_plot_stats):
                        plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder,model = model_name, portion = "all")
                
                elif(testing_portion == "subset"):
                    # iterate through files and record predictions
                    rows, time_taken, duration = sm_predict_subset_excel(frequency= model_folder[:3], detector = detector, emotions = emotions, rows = 2, cols = 1, sheet = sheet, time_taken = [], duration = [])
                    rows, time_taken, duration = predict_subset_excel(frequency = model_folder[:3], detector = detector, folder = "JL", rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)
                    rows, time_taken, duration = predict_subset_excel(frequency = model_folder[:3], detector = detector, folder = "Nene", rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)

                    # save and display where the predictions are saved
                    wb.save('test_audio/prediction.xlsx')
                    print('predictions saved to test_audio/prediction.xlsx')

                    # if plotting is selected then display plot, otherwise predict Nene audio - Can't get duration of nene audio as not in float form
                    if (is_plot_stats):
                        plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder,model = model_name, portion = "subset")

                # if not subset or all, then exit
                else:
                    print("Please select all or subset in Testing portion in TEST MULTIPLE SETTING")
                    sys.exit("Please select all or subset in Testing portion in TEST MULTIPLE SETTING")
                
            # else if prediction_output_modeting to text
            elif(prediction_output_mode == "text"):

                with open(file = os.path.join('test_audio','predictions.txt'), mode  = 'w') as text_file:
                    text_file.write("results," + str(emotions) + "\n")

                    if(testing_portion == "all"):

                        time_taken, duration = predict_all_text(detector = detector ,csv_file = 'audio_csv/test_emodb.csv', text_file = text_file, time_taken = [], duration = [])
                        time_taken, duration = predict_all_text(detector = detector ,csv_file = 'audio_csv/test_tess_ravdess.csv', text_file = text_file, time_taken =time_taken, duration = duration)
                        time_taken, duration = predict_all_text(detector = detector ,csv_file = 'audio_csv/test_custom.csv', text_file = text_file, time_taken =time_taken, duration = duration)
                        print('predictions saved to test_audio/prediction.txt')

                        # if plotting the time taken is selected, plot the time taken to predict with the length of the audio
                        if (is_plot_stats):
                            plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder,model = model_name, portion = "all")
                    
                    elif(testing_portion == "subset"):
                        # iterate through files and record predictions
                        time_taken, duration = sm_predict_subset_text(frequency= model_folder, detector = detector, emotions = emotions, file = text_file, time_taken=[], duration = [])
                        time_taken, duration = predict_subset_text(frequency= model_folder, detector = detector, file = text_file, folder = "JL", time_taken = time_taken, duration = duration)
                        time_taken, duration = predict_subset_text(frequency= model_folder, detector = detector, file = text_file, folder = "Nene", time_taken = time_taken, duration = duration)
                        print('predictions saved to test_audio/prediction.txt')

                        if (is_plot_stats):
                            plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder,model = model_name, portion = "subset")

                    # if not subset or all, then exit
                    else:
                        print("Please select all or subset in Testing portion in TEST MULTIPLE SETTING")
                        sys.exit("Please select all or subset in Testing portion in TEST MULTIPLE SETTING")

        # if predicting live audio
        elif(mode == 'live'):

            # record
            print("Please talk")
            buffer, sample_width  = record_without_file()

            # predict
            start_predict = time.perf_counter()
            emotion_probabilities = detector.predict_proba_audio_buffer(buffer)
            end_predict = time.perf_counter()

            # display predictions
            print(f"\n emotion probabilities \n{emotion_probabilities}")

            # get most likely emotion
            most_likely_emotion_key = max(emotion_probabilities, key=emotion_probabilities.get)
            most_likely_emotion_value = emotion_probabilities[most_likely_emotion_key]
            del emotion_probabilities[most_likely_emotion_key]

            # get second most likely emotion
            second_likely_emotion_key = max(emotion_probabilities, key=emotion_probabilities.get)
            second_likely_emotion_value = emotion_probabilities[second_likely_emotion_key]

            effective_prob = (most_likely_emotion_value - second_likely_emotion_value)*100
            # print to user 
            print(f"\nBest prediction  : {most_likely_emotion_key} \nsecond best prediction : {second_likely_emotion_key} \ndifference is {effective_prob} %")
            print(f"\nTherefore predicted {most_likely_emotion_key} with an effective probability of {effective_prob} %")
            print(f"Time it took to predict: {(end_predict - start_predict)*1000} ms")

        elif(mode == 'confusion'):
            detector.load_data()
            print(detector.confusion_matrix())
            print("Test accuracy score: {:.3f}%".format(detector.test_score()*100))

        # if single or multiple is not chosen
        else:
            print("Please choose whether to predict single, multiple, live, or confusion.\n This can be done under Test Settings, Test mode in predict.json")
            sys.exit("Please choose whether to predict single,multiple, live or confusion.\n This can be done under Test Settings, Test mode in predict.json")
    
    # if train
    if(test_or_train == "train"):
        # if train mode is multiple, get the list of classifiers that were entered
        if(mode == "multiple"):
            model_to_train = mode_settings['Multiple classifiers to train']
        
        # if single, get the classifier that was chosen in the mandatory settings
        elif(mode == "single"):
            model_to_train = [model_name]
        
        else:
            print("Please choose whether to predict single or multiple.\n This can be done under Test Settings, Test mode in predict.json")
            sys.exit("Please choose whether to predict single or multiple.\n This can be done under Test Settings, Test mode in predict.json")
        
        # get previously tuned models 
        pre_tuned_models = get_pre_tuned_models(True)                                               # gets pre - grid tuned models run gridsearch.py for MLP classifier, others were already included in repo
        pre_tuned_model_name, pre_tuned_models_dict = get_pre_tuned_models_dict(pre_tuned_models)   # Formats these pre - grid tuned models as dictionaries
            
        # start timing in order to display the time taken to train the model(s) later on 
        start_train = time.perf_counter()


        # join model name
        for model in model_to_train:

            # navigate to the directory of the models in the list
            model_dir = os.path.join(model_folder,model)

            #load model parameters that are saved as a dictionary returned in grid_tuned_models_dict()
            if(model == "MLPClassifier"):
                detector = EmotionRecognizer(pre_tuned_models_dict[model.format(pre_tuned_model_name)] , emotions=emotions, model_dir = model_dir, features=features, verbose=0)
            else:
                detector = EmotionRecognizer(emotions=emotions, model_dir = model_dir, features=features, verbose=0)
            # train the model and display status, and print the confusion matrix
            detector.train()
            print(f"\n{model} trained")
            print(detector.confusion_matrix())
            print("Test accuracy score: {:.3f}%".format(detector.test_score()*100))

        # display how long the whole process took 
        end_train = time.perf_counter()
        print(f"\nThis process took {end_train - start_train} seconds")
    


                

    