import os
import librosa
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import time
from csv import reader
import statistics
import numpy as np
import seaborn as sns
import math

def sm_predict_subset_excel(frequency, detector, emotions, rows, cols, sheet, time_taken, duration):
    """ Predicts a subset of sm audio and outputs the predictions to an excel spreadsheet 

        Parameters:
        -----------
        frequency   : Sampling rate of the audio file
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        emotions    : emotions to predict
        rows        : row of the spreadsheet to start from
        cols        : column of the spreadsheet to start from
        sheet       : spreadsheet to output to
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    sheet[get_column_letter(cols) + str(rows)] = "sm audio"
    rows += 1

    for emotion in emotions:
        for audio in os.listdir(os.path.join("test_audio",f"emotion testing audio {frequency}",f"{emotion}")):

            # calculate probability
            duration.append(librosa.get_duration(filename = os.path.join("test_audio",f"emotion testing audio {frequency}",emotion, audio)))

            start_predict = time.perf_counter()
            predictions = detector.predict_proba_file(os.path.join("test_audio",f"emotion testing audio {frequency}",emotion, audio))
            end_predict = time.perf_counter()

            time_taken.append((end_predict - start_predict)*1000)
            
            sheet[get_column_letter(cols) + str(rows)] = str(emotion)
            _, intensity, _ = audio.split("_")
            sheet[get_column_letter(cols) + str(rows)] = str(emotion)
            sheet[get_column_letter(cols + 1) + str(rows)] = intensity
            cols += 2 

            # record if the perediction is correct
            if(emotion==max(predictions, key=predictions.get).lower()):
                sheet[get_column_letter(cols) + str(rows)] = "correct"
                cols += 1        
            else:
                sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()}"
                cols += 1


            # record the result
            for value in (predictions).values():
                sheet[get_column_letter(cols) + str(rows)] = value
                cols += 1

            rows += 1
            cols = 1
    
    return(rows, time_taken, duration)

def predict_subset_excel(frequency, detector, folder, rows, cols, sheet, time_taken, duration):
    """ Predicts audio files in a sub folder in test_audio folder and outputs the predictions to an excel spreadsheet 

        Parameters:
        -----------
        frequency   : Sampling rate of the audio file
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        folder      : folder in test_audio that audio files to predict are stored in
        rows        : row of the spreadsheet to start from
        cols        : column of the spreadsheet to start from
        sheet       : spreadsheet to output to
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    sheet[get_column_letter(cols) + str(rows)] = f"{folder} audio"
    rows += 1

    for audio in os.listdir(os.path.join('test_audio',f'{folder}_{frequency}')):
        sentiment = audio.split("_")

        # Get prediction and record the correct sentiment
        duration.append(librosa.get_duration(filename = os.path.join('test_audio',f'{folder}_{frequency}',audio)))

        start_predict = time.perf_counter()
        predictions = detector.predict_proba_file(os.path.join('test_audio',f'{folder}_{frequency}',audio))
        end_predict = time.perf_counter()

        time_taken.append((end_predict - start_predict)*1000)
        
        sheet[get_column_letter(cols) + str(rows)] = str(sentiment[1])
        sheet[get_column_letter(cols + 1) + str(rows)] = str(sentiment[2])

        cols += 2

        # record if the perediction is correct
        if(sentiment[1]==max(predictions, key=predictions.get).lower()):
            sheet[get_column_letter(cols) + str(rows)] = "correct"
            cols += 1        
        else:
            sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()}"
            cols += 1

        # record the result
        for value in (predictions).values():
            sheet[get_column_letter(cols) + str(rows)] = value
            cols += 1

        rows += 1
        cols = 1
    
    return(rows, time_taken, duration)

def predict_all_excel(detector, rows, cols, sheet, file, time_taken, duration):
    """ Predicts all custom testing audio specified in a csv file (test_custom.csv, test_tess_ravdess.csv, test_emodb.csv) and outputs the predictions to an excel spreadsheet 

        Parameters:
        -----------
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        rows        : row of the spreadsheet to start from
        cols        : column of the spreadsheet to start from
        sheet       : spreadsheet to output to
        file        : csv file where the testing audio is recorded, choose from test_tess_ravdess.csv, test_emodb.csv, test_custom.csv
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    sheet[get_column_letter(cols) + str(rows)] = file
    rows += 1

    with open(file, 'r') as audio_file:
        csv_reader = reader(audio_file)
        header = next(csv_reader)

        if header != None:
            for row in csv_reader:
                audio = row[1]
                emotion = row[2]

                duration.append(librosa.get_duration(filename = audio))

                start_predict = time.perf_counter()
                predictions = detector.predict_proba_file(audio)
                end_predict = time.perf_counter()

                time_taken.append((end_predict - start_predict)*1000)

                # record correct emotion
                sheet[get_column_letter(cols) + str(rows)] = emotion

                cols += 2

                # record if the perediction is correct
                if(emotion==max(predictions, key=predictions.get).lower()):
                    sheet[get_column_letter(cols) + str(rows)] = "correct"
                    cols += 1        
                else:
                    sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()}"
                    cols += 1
                
                for value in (predictions).values():
                    sheet[get_column_letter(cols) + str(rows)] = value
                    cols += 1
                sheet[get_column_letter(cols) + str(rows)] = audio

                rows += 1
                cols = 1

    return(rows, time_taken, duration)

def sm_predict_subset_text(frequency, detector, emotions, file, time_taken, duration):
    """ Predicts a subset of sm audio and outputs the predictions to a specified text file

        Parameters:
        -----------
        frequency   : Sampling rate of the audio file
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        emotions    : emotions to predict
        file        : text file to output to
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    # iterate through all the files

    for emotion in emotions:

        # record emotions to predict to write to the putput file later
        emotion_to_predict = "\n" + emotion + (8-len(emotion))*(" ")

        for audio in os.listdir(os.path.join("test_audio",f"emotion testing audio {frequency}",f"{emotion}")):   
            # write if prediction was correct
            duration.append(librosa.get_duration(filename = audio))

            start_predict = time.perf_counter()
            predictions = detector.predict_proba_file(os.path.join("test_audio",f"emotion testing audio {frequency}",emotion, audio))
            end_predict = time.perf_counter()

            time_taken.append((end_predict - start_predict)*1000)

            observed_emotion = max(predictions, key=predictions.get).lower()
            if(emotion==observed_emotion):
                file.write(emotion_to_predict + " correct           :" )
            else:
                wrong = str(observed_emotion) + (8-len(observed_emotion))*(" ")
                file.write(f"{emotion_to_predict} incorrect {wrong}:")
            
            # Write probabiltiy distribution
            for value in predictions.values():
                file.write('%.2f' % value)
                file.write(',')
    return(time_taken, duration)

def predict_subset_text(frequency, detector, file, folder, time_taken, duration):
    """ Predicts a subset of sm audio and outputs the predictions to a specified text file

        Parameters:
        -----------
        frequency   : Sampling rate of the audio file
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        emotions    : emotions to predict
        file        : text file to output to
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    # iterate through all the files
    audio_folder = os.path.join('test_audio',f'{folder}_{frequency[:3]}')
    file.write(f'\n\nPredicting {folder} audio files')
    for audio in os.listdir(audio_folder):  
        sentiment = audio.split("_")
        emotion_to_predict = "\n" + sentiment[1] + (8-len(sentiment[1]))*(" ")

        # write if prediction was correct
        duration.append(librosa.get_duration(filename = os.path.join('test_audio',f'{folder}_{frequency[:3]}',audio)))

        start_predict = time.perf_counter()
        predictions = detector.predict_proba_file(os.path.join('test_audio',f'{folder}_{frequency[:3]}',audio))
        end_predict = time.perf_counter()

        time_taken.append((end_predict - start_predict)*1000)

        observed_emotion = max(predictions, key=predictions.get).lower()
        if(sentiment[1]==observed_emotion):
            file.write(emotion_to_predict + " correct           :" )
        else:
            wrong = str(observed_emotion) + (8-len(observed_emotion))*(" ")
            file.write(f"{emotion_to_predict} incorrect {wrong}:")
        
        # Write probabiltiy distribution
        for value in predictions.values():
            file.write('%.2f' % value)
            file.write(',')
    return(time_taken, duration)

def predict_all_text(detector,csv_file, text_file, time_taken, duration):
    """ Predicts all custom testing audio specified in a csv file (test_custom.csv, test_tess_ravdess.csv, test_emodb.csv) and outputs the predictions to an excel spreadsheet 

        Parameters:
        -----------
        detector    : The instance of EmotionalRecognzer used to predict sentiment
        rows        : row of the spreadsheet to start from
        cols        : column of the spreadsheet to start from
        sheet       : spreadsheet to output to
        file        : csv file where the testing audio is recorded, choose from test_tess_ravdess.csv, test_emodb.csv, test_custom.csv
        time_taken  : list that stores the time taken for each prediction
        duration    : list that stores the duration/length of each audio file

        Returns:
        --------
        rows        : next free row after writing predictions
        time_taken  : list that stores the time taken for each prediction, with times taken added
        duration    : list that stores the duration/length of each audio file, with length of files added

    """
    text_file.write(f'\n\nPredicting {csv_file} audio files')

    with open(csv_file, 'r') as audio_file:
        csv_reader = reader(audio_file)
        header = next(csv_reader)

        if header != None:
            for row in csv_reader:
                audio = row[1]
                emotion_to_predict = "\n" + row[2] + (8-len(row[2]))*(" ")

                duration.append(librosa.get_duration(filename = audio))

                start_predict = time.perf_counter()
                predictions = detector.predict_proba_file(audio)
                end_predict = time.perf_counter()

                time_taken.append((end_predict - start_predict)*1000)

                observed_emotion = max(predictions, key=predictions.get).lower()
                if(row[2]==observed_emotion):
                    text_file.write(emotion_to_predict + " correct           :" )
                else:
                    wrong = str(observed_emotion) + (8-len(observed_emotion))*(" ")
                    text_file.write(f"{emotion_to_predict} incorrect {wrong}:")
                
                # Write probabiltiy distribution
                for value in predictions.values():
                    text_file.write('%.2f' % value)
                    text_file.write(',')

    return(time_taken, duration)

def plot_time_taken(duration, time_taken, frequency, model, portion):
    """ Takes duration of audio files predicted and time taken to predict, and outputs them as a visualisation
        Saves plot as an png file

        Parameters:
        -----------
        duration    : list that stores the duration/length of each audio file
        time_taken  : list that stores the time taken for each prediction
        frequency   : frequency_features used to predict eg 16k_3feat
        model       : model used to predict
        portion     : portion of the dataset it was tested on 

    """
    median_time = statistics.median(time_taken)
    mean_time = statistics.mean(time_taken)

    median_length = statistics.median(duration)
    mean_length = statistics.mean(duration)

    # sort sample length and in ascending order of audio length
    index = np.argsort(duration)
    sorted_duration = np.zeros(len(duration))
    sorted_time_taken = np.zeros(len(index))

    for i in range(len(index)):
        sorted_time_taken[i] = time_taken[index[i]]
        sorted_duration[i] = duration[index[i]]

    # create subplot axes to plot on 
    gs = dict(width_ratios=[1, 1], height_ratios=[1, 6, 6, 12])
    fig, axd = plt.subplot_mosaic([['top', 'top'],
                                ['upper left', 'right'],
                               ['mid left', 'right'],
                               ['lower','lower']],
                              gridspec_kw=gs,
                              figsize = (15,8),
                              dpi = 200,
                              constrained_layout=True)
    ax0 = axd['top']
    ax0.axis("off")
    ax0.text(0, 0, f'number of samples:{len(duration)}', style='italic', fontsize = 12,
         bbox={'facecolor': 'grey', 'alpha': 0.5, 'pad': 5})
    
    # plot histogram and probability density of the time taken to predict
    ax1  = axd['upper left']
    sns.set(style = 'ticks')
    hist1 = sns.histplot(data = time_taken, kde = True, bins = 20,edgecolor = 'lightsteelblue', ax = ax1)
    start, end = hist1.get_ylim()
    if (end <= 10):
        hist1.set_yticks(np.arange(start,end,1))
    start, end = hist1.get_xlim()
    hist1.set_xticks(np.arange(0,end,5))


    ax1.axvline(x = median_time, color = 'g' , label = f"median: {round(median_time,2)} ms")
    ax1.axvline(x = mean_time, color = 'm', label = f"mean: {round(mean_time,2)} ms")
    ax1.set_xlabel("Time taken to predict (ms)", fontsize = 12)
    ax1.set_ylabel("Audio count", fontsize = 12)
    ax1.legend(loc="upper right")

    # plot histogram and probability density for the duration of audio
    ax2 = axd['mid left']
    hist2 = sns.histplot(x = duration, kde = True, bins = 20, edgecolor = 'lightsteelblue', ax = ax2)
    start, end = hist2.get_ylim()
    if (end <= 10):
        hist2.set_yticks(np.arange(start,end,1))
    start, end = hist2.get_xlim()
    hist2.set_xticks(np.arange(0,end,1))

    ax2.axvline(x = median_length, color = 'g', label = f"median: {round(median_length,2)} s")
    ax2.axvline(x = mean_length, color = 'm', label = f"mean: {round(mean_length,2)} s")
    ax2.set_xlabel("Duration Of Audio (s)", fontsize = 12)
    ax2.set_ylabel("Audio count", fontsize = 12)
    ax2.legend(loc="upper right")

    # plot time it took to predict and the length of the audio together, in ascending order of the length of audio
    ax3 = axd['lower']
    ax4 = ax3.twinx()
    lns1 = ax3.plot(list(range(1,len(time_taken)+1)), sorted_time_taken, '-b.', label = "Time taken to predict audio")
    lns2 = ax4.plot(list(range(1,len(time_taken)+1)), sorted_duration, '-r.', label = "Duration of audio")

    ax4.set_ylabel("Duration Of Audio (s)", fontsize = 12)
    ax3.set_xlabel("Audio Samples", fontsize = 12)
    ax3.set_ylabel("Time Taken To Predict Audio (ms)", fontsize = 12)
    lns = lns1 + lns2
    labs = [l.get_label() for l in lns]
    ax3.legend(lns, labs, loc="upper left")
    ax3.grid(True)

    # plot a scatter plot of the time taken to predict and the length of the corresponding audio
    ax5 = axd['right']
    ax5.plot(time_taken, duration, 'r.')
    ax5.set_xlabel('Time Taken To Predict Audio (ms)', fontsize = 12)
    ax5.set_ylabel('Duration Of Audio (s)', fontsize = 12)
    ax5.grid(True)

    # add a title and lplot
    #Prediction time(s) for 16KHz speech audio files using the MLP classifier
    fig.suptitle(f"Prediction time(s) for {frequency[:3]}Hz speech audio using {model}", fontsize = 21)
    
    plt.tight_layout()
    plt.savefig(f'performance_plots/{frequency}_{model}_{portion}.png')
    print(f'Plot of time taken to predict saved to performance_plots/{frequency}_{model}_{portion}.png' )

def get_long_audio(files):
    with open(os.path.join("test_audio","long_audio.txt"), "w") as long_audio_file:
        for file in files:
            with open(file, 'r') as audio_file:
                csv_reader = reader(audio_file)
                header = next(csv_reader)

                if header != None:
                    for row in csv_reader:
                        audio = row[1]
                        audio_duration = librosa.get_duration(filename = audio)

                        if audio_duration > 15:
                            long_audio_file.write(f"{audio_duration},{audio}\n")

