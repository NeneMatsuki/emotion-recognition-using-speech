from openpyxl.reader import excel
from emotion_recognition import EmotionRecognizer
from deep_emotion_recognition import DeepEmotionRecognizer
import os
import glob
import json
from array import array
from struct import pack
from sklearn.ensemble import GradientBoostingClassifier, BaggingClassifier, RandomForestClassifier, AdaBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import time
from csv import reader

def sm_predict_excel(frequency, detector, emotions, rows, cols, sheet):
    sheet[get_column_letter(cols) + str(rows)] = "sm audio"
    rows += 1

    for emotion in emotions:
        for audio in os.listdir(os.path.join("predict_from_audio",f"emotion testing audio {frequency}",f"{emotion}")):

            # calculate probability
            predictions = detector.predict_proba(os.path.join("predict_from_audio",f"emotion testing audio {frequency}",emotion, audio))
            sheet[get_column_letter(cols) + str(rows)] = str(emotion)
            _, intensity, _ = audio.split("_")
            sheet[get_column_letter(cols) + str(rows)] = str(emotion)
            sheet[get_column_letter(cols + 1) + str(rows)] = intensity
            cols += 2 

            # record if the perediction is correct
            if(emotion==max(predictions, key=predictions.get).lower()):
                sheet[get_column_letter(cols) + str(rows)] = "correct"
                cols += 1        
            else:
                sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()} : "
                cols += 1


            # record the result
            for value in (predictions).values():
                sheet[get_column_letter(cols) + str(rows)] = value
                cols += 1

            rows += 1
            cols = 1
    
    return(rows)

def predict_excel(frequency, detector, folder, rows, cols, sheet):
    sheet[get_column_letter(cols) + str(rows)] = f"{folder} audio"
    rows += 1

    for audio in os.listdir(os.path.join('predict_from_audio',f'{folder}_{frequency}')):
        sentiment = audio.split("_")

        # Get prediction and record the correct sentiment
        predictions = detector.predict_proba(os.path.join('predict_from_audio',f'{folder}_{frequency}',audio))
        sheet[get_column_letter(cols) + str(rows)] = str(sentiment[1])
        sheet[get_column_letter(cols + 1) + str(rows)] = str(sentiment[2])

        cols += 2

        # record if the perediction is correct
        if(sentiment[1]==max(predictions, key=predictions.get).lower()):
            sheet[get_column_letter(cols) + str(rows)] = "correct"
            cols += 1        
        else:
            sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()}"
            cols += 1

        # record the result
        for value in (predictions).values():
            sheet[get_column_letter(cols) + str(rows)] = value
            cols += 1

        rows += 1
        cols = 1
    
    return(rows)

def sm_predict_text(frequency, detector, emotions, file):
        # iterate through all the files
        file.write("results," + str(emotions) + "\n")
        for emotion in emotions:

            # record emotions to predict to write to the putput file later
            to_predict = "\n" + emotion + (8-len(emotion))*(" ")

            for audio in os.listdir(os.path.join("predict_from_audio",f"emotion testing audio {frequency}",f"{emotion}")):   
                # write if prediction was correct
                predictions = detector.predict_proba(os.path.join("predict_from_audio",f"emotion testing audio {frequency}",emotion, audio))

                observed_emotion = max(predictions, key=predictions.get).lower()
                if(emotion==observed_emotion):
                    file.write(to_predict + " correct           :" )
                else:
                    wrong = str(observed_emotion) + (8-len(observed_emotion))*(" ")
                    file.write(f"{to_predict} incorrect {wrong}:")
                
                # Write probabiltiy distribution
                for value in predictions.values():
                    file.write('%.2f' % value)
                    file.write(',')
                
def sm_predict_all_excel(detector, rows, cols, sheet):

    sheet[get_column_letter(cols) + str(rows)] = "sm audio"
    rows += 1


    with open('test_custom.csv', 'r') as test_file:
        csv_reader = reader(test_file)
        header = next(csv_reader)

        if header != None:
            for row in csv_reader:
                audio = row[1]
                emotion = row[2]

                predictions = detector.predict_proba(audio)

                # record correct emotion
                sheet[get_column_letter(cols) + str(rows)] = emotion

                # get intensity if applicable
                audio_info = audio.split("_")
                if ((audio_info[1] == "high") or (audio_info[1] == "med") or (audio_info[1] == "low")):
                    sheet[get_column_letter(cols + 1) + str(rows)] = audio_info[1]
                cols += 2

                # record if the perediction is correct
                if(emotion==max(predictions, key=predictions.get).lower()):
                    sheet[get_column_letter(cols) + str(rows)] = "correct"
                    cols += 1        
                else:
                    sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()} : "
                    cols += 1
                
                for value in (predictions).values():
                    sheet[get_column_letter(cols) + str(rows)] = value
                    cols += 1

                rows += 1
                cols = 1

    return(rows)

def predict_all_excel(detector, rows, cols, sheet, file):
    sheet[get_column_letter(cols) + str(rows)] = file
    rows += 1

    with open(file, 'r') as test_file:
        csv_reader = reader(test_file)
        header = next(csv_reader)

        if header != None:
            for row in csv_reader:
                audio = row[1]
                emotion = row[2]

                predictions = detector.predict_proba(audio)

                # record correct emotion
                sheet[get_column_letter(cols) + str(rows)] = emotion

                cols += 2

                # record if the perediction is correct
                if(emotion==max(predictions, key=predictions.get).lower()):
                    sheet[get_column_letter(cols) + str(rows)] = "correct"
                    cols += 1        
                else:
                    sheet[get_column_letter(cols) + str(rows)] = f"incorrect {max(predictions, key=predictions.get).lower()}"
                    cols += 1
                
                for value in (predictions).values():
                    sheet[get_column_letter(cols) + str(rows)] = value
                    cols += 1

                rows += 1
                cols = 1

    return(rows)

        
