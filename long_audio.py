from emotion_recognition import EmotionRecognizer
from emotion_recognizer_functions import predict_all_excel
from utils import get_grid_tuned_models, load_mandatory_settings, load_testing_settings, load_training_settings
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from csv import reader
import librosa
import time
from pydub import AudioSegment

def get_grid_tuned_models_dict(estimators):
    result = [ '"{}"'.format(estimator.__class__.__name__) for estimator, _, _ in estimators ]
    return ','.join(result), {estimator_name.strip('"'): estimator for estimator_name, (estimator, _, _) in zip(result, estimators)}

if __name__ == "__main__":
    # load and check if json file exists
    json_file = "test_train_config.json"
    if(not os.path.isfile(json_file)):
        print(f"config file, {json_file} does not exist")
        sys.exit(f"config file, {json_file} does not exist")

    # get model parameters that are grid tuned
    grid_tuned_models = get_grid_tuned_models(True)
    grid_tuned_models_name, grid_tuned_models_dict = get_grid_tuned_models_dict(grid_tuned_models)
    Test_or_train_mode, classifier_name, model_folder, emotions, features, model_dir = load_mandatory_settings(json_file)
    
    detector = EmotionRecognizer(grid_tuned_models_dict[classifier_name] , emotions=emotions, model_dir = model_dir, features=features, verbose=0)
    # initialise workbook 
    wb = load_workbook('predict_from_audio/long_audio.xlsx')
    wb.remove(wb['long_audio'])
    wb.create_sheet('long_audio')
    sheet = wb['long_audio']
    sheet["A1"] = "Audio length"
    sheet["B1"] = "True emotion"
    sheet["C1"] = "Predicted emotion"
    sheet["D1"] = "Time taken to predict" 
    
    rows = 2
    cols = 1
    time_taken = []
    duration = []
    
    file = 'long_audio.csv'

    with open(file, 'r') as audio_file:
        csv_reader = reader(audio_file)
        header = next(csv_reader)

        if header != None:
            for row in csv_reader:
                audio = row[1]
                emotion = row[2]
                audio_length = librosa.get_duration(filename = audio)
                duration.append(audio_length)

                sheet[get_column_letter(cols) + str(rows)] = str(audio_length)
                cols += 1

                start_predict = time.perf_counter()
                predictions = detector.predict_proba(audio)
                end_predict = time.perf_counter()

                prediction_time = (end_predict - start_predict)*1000
                time_taken.append(prediction_time)

                # record correct emotion
                sheet[get_column_letter(cols) + str(rows)] = emotion
                sheet[get_column_letter(cols + 1) + str(rows)] = max(predictions, key=predictions.get)
                sheet[get_column_letter(cols + 2) + str(rows)] = prediction_time
                cols += 4


                audio_length = audio_length*1000
                chunk_length = 5000
                segments = int(audio_length/chunk_length)

                for i in range(segments):
                    new_audio = AudioSegment.from_wav(audio)
                    new_audio = new_audio[i*chunk_length:(i+1)*chunk_length]
                    new_audio.export('temp.wav', format="wav")

                    start_predict = time.perf_counter()
                    predictions = detector.predict_proba('temp.wav')
                    end_predict = time.perf_counter()
                    sheet[get_column_letter(cols) + str(rows)] = max(predictions, key=predictions.get)
                    sheet[get_column_letter(cols + 1) + str(rows)] = (end_predict - start_predict)*1000
                    cols += 2

                new_audio = AudioSegment.from_wav(audio)
                new_audio = new_audio[i*chunk_length:]
                new_audio.export('temp.wav', format="wav")

                start_predict = time.perf_counter()
                predictions = detector.predict_proba('temp.wav')
                end_predict = time.perf_counter()
                sheet[get_column_letter(cols) + str(rows)] = max(predictions, key=predictions.get)
                sheet[get_column_letter(cols + 1) + str(rows)] = (end_predict - start_predict)*1000
                cols += 2

                rows += 1
                cols = 1

    wb.save('predict_from_audio/long_audio.xlsx')