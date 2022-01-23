from emotion_recognition import EmotionRecognizer
import os
import sys
from sklearn.tree import DecisionTreeClassifier
import time
import librosa
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from emotion_recognizer_functions import *

from utils import get_grid_tuned_models, load_mandatory_settings, load_testing_settings, load_training_settings
from live_voice_emotion_recognizer import record_without_file

def get_grid_tuned_models_dict(estimators):
    result = [ '"{}"'.format(estimator.__class__.__name__) for estimator, _, _ in estimators ]
    return ','.join(result), {estimator_name.strip('"'): estimator for estimator_name, (estimator, _, _) in zip(result, estimators)}

if __name__ == "__main__":

    # load and check if json file exists
    json_file = "test_train_config.json"
    if(not os.path.isfile(json_file)):
        print(f"config file, {json_file} does not exist")
        sys.exit(f"config file, {json_file} does not exist")

    # get model parameters that are grid tuned
    grid_tuned_models = get_grid_tuned_models(True)
    grid_tuned_models_name, grid_tuned_models_dict = get_grid_tuned_models_dict(grid_tuned_models)
    Test_or_train_mode, classifier_name, model_folder, emotions, features, model_dir = load_mandatory_settings(json_file)

    # If testing
    if(Test_or_train_mode == "test"):

        # load testing parameters
        test_mode, audio_or_output, plot_time = load_testing_settings(json_file)

        # create detector instance
        detector = EmotionRecognizer(grid_tuned_models_dict[classifier_name] , emotions=emotions, model_dir = model_dir, features=features, verbose=0)

        # if predicting a single audio
        if(test_mode == 'single'):

            # collect the audio directory
            audio_dir = audio_or_output

            # print to user what the settings are 
            print(f'\nChosen to test a single audio using {classifier_name} trained on {model_folder}')
            print(f'Length of audio: {librosa.get_duration(filename = audio_dir)} seconds')

            #predict from filename passed in args
            start_predict = time.perf_counter()
            result = detector.predict_proba(audio_dir)
            end_predict = time.perf_counter()

            # print results 
            print(f'\n{result}')
            maximum = max(result, key=result.get)
            max_value = result[maximum]
            del result[maximum]
            second = max(result, key=result.get)
            second_value = result[second]

            print(f"\nMost likely to be  : {maximum} \nsecond most likely to be : {second} \ndifference is {(max_value - second_value)*100} %")
            print(f"\nTime it took to predict: {(end_predict - start_predict)*1000} ms")

        # if predicting multiple audio
        elif(test_mode == 'multiple'):

            # get prefferend methpod of displaying predicitons   
            display_predictions = audio_or_output

            # if display is excel and predicting all 10521 testing dataset
            if(display_predictions == "excel all"):
                # initialise workbook 
                wb = load_workbook('predict_from_audio/prediction.xlsx')
                wb.remove(wb['predictions'])
                wb.create_sheet('predictions')
                sheet = wb['predictions']
                sheet["A1"] = "True emotion"
                sheet["B1"] = "Intensity" 
                sheet["C1"] = "Result"

                # add labels to the workbook
                for i in range(len(emotions)):
                    sheet[get_column_letter(i + 4) + "1"] =  emotions[i]
                
                sheet[get_column_letter(i + 4) + "1"] = "audio file"
                rows = 2

                # array to record time taken and duration
                time_taken = []
                duration = []

                # read through all files, and record the predicitons
                rows, time_taken, duration = sm_predict_all_excel(detector = detector,rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)
                rows, time_taken, duration = predict_all_excel(detector = detector,rows = rows, cols = 1, sheet = sheet, file = 'test_tess_ravdess.csv', time_taken = time_taken, duration = duration)
                rows, time_taken, duration = predict_all_excel(detector = detector,rows = rows, cols = 1, sheet = sheet, file = 'test_emodb.csv', time_taken = time_taken, duration = duration)                    

                # save and display where the predictions ar saved
                wb.save('predict_from_audio/prediction.xlsx')
                print('predictions saved to predict_from_audio/prediction.xlsx')

                # if plotting the time taken is selected, plot the time taken to predict with the length of the audio
                if (plot_time.lower() == 'yes'):
                    plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder,model = classifier_name)

            # if testing subset of testing file (71 files)
            elif(display_predictions.lower() == 'excel subset'):
                # initialise workbook 
                wb = load_workbook('predict_from_audio/prediction.xlsx')
                wb.remove(wb['predictions'])
                wb.create_sheet('predictions')
                sheet = wb['predictions']
                sheet["A1"] = "True emotion"
                sheet["B1"] = "Intensity" 
                sheet["C1"] = "Result"

                # record emotions
                for i in range(len(emotions)):
                    sheet[get_column_letter(i + 4) + "1"] =  emotions[i]
                
                rows = 2
                time_taken = []
                duration = []

                # iterate through files and record predictions
                rows, time_taken, duration = sm_predict_excel(frequency= model_folder[:3], detector = detector, emotions = emotions, rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)
                rows, time_taken, duration = predict_excel(frequency = model_folder[:3], detector = detector, folder = "Nene", rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)
                rows, time_taken, duration = predict_excel(frequency = model_folder[:3], detector = detector, folder = "JL", rows = rows, cols = 1, sheet = sheet, time_taken = time_taken, duration = duration)

                # save and display where the predictions are saved
                wb.save('predict_from_audio/prediction.xlsx')
                print('predictions saved to predict_from_audio/prediction.xlsx')

                # if plotting is selected then display plot
                if (plot_time == 'yes'):
                    plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder[:3],model = classifier_name)

            # else if display_predictionsting to text
            else:
                print("This is not implemented fully, only can predict from sm subset audio")

                with open(file = 'predict_from_audio' + os.sep + 'predictions.txt', mode  = 'w') as file:
                    time_taken, duration = sm_predict_text(frequency= model_folder, detector = detector, emotions = emotions, file = file)
                    print('predictions saved to predict_from_audio/prediction.txt')
                if (plot_time == 'yes'):
                    plot_time_taken(duration = duration, time_taken = time_taken, frequency = model_folder[:3],model = classifier_name)
        
        # if predicting live audio
        elif(test_mode == 'live'):

            # record
            print("Please talk")
            data, sample_width  = record_without_file()

            # predict
            start_predict = time.perf_counter()
            result = detector.predict_proba_audio(data)
            end_predict = time.perf_counter()

            # display predictions
            print(f"\n emotion probabilities \n{result}")

            maximum = max(result, key=result.get)
            max_value = result[maximum]
            del result[maximum]

            second = max(result, key=result.get)
            second_value = result[second]

            print(f"\nmost likely to be  : {maximum} \nsecond most likely to be : {second} \ndifference is {(max_value - second_value)*100} %")
            print(f"\nTime it took to predict: {(end_predict - start_predict)*1000}ms")

        # if single or multiple is not chosen
        else:
            sys.exit("Please choose whether to predict single or multiple.\n This can be done under Testing Settings, Test mode in predict.json")
    
    # if training
    else:
        # get training settings
        train_mode, classifier_list = load_training_settings(json_file)

        # if training mode is multiple, get the list of classifiers that were entered
        if(train_mode == "multiple"):
            train_classifiers = classifier_list
        
        # if single, get the classifier that was chosen in the mandatory settings
        else:
            train_classifiers = [classifier_name]

        # start timing in order to display the time taken to train the model(s) later on 
        start_train = time.perf_counter()

        # join model name
        for model in train_classifiers:

            # navigate to the directory of the models in the list
            model_dir = os.path.join(model_folder,model)

            #load model parameters that are saved as a dictionary returned in grid_tuned_models_dict()

            detector = EmotionRecognizer(grid_tuned_models_dict[model.format()] , emotions=emotions, model_dir = model_dir, features=features, verbose=0)
            
            # train the model and display status, and print the confusioni matrix
            detector.train()
            print(f"\n{model} trained")
            print(detector.confusion_matrix())
            print("Test accuracy score: {:.3f}%".format(detector.test_score()*100))

        # display how long the whole process took 
        end_train = time.perf_counter()
        print(f"\nThis process took {end_train - start_train} seconds")
    


                

    