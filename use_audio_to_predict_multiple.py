from emotion_recognition import EmotionRecognizer
from deep_emotion_recognition import DeepEmotionRecognizer
import os
import glob
import json
from array import array
from struct import pack
from sklearn.ensemble import GradientBoostingClassifier, BaggingClassifier, RandomForestClassifier, AdaBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import time

from utils import get_best_estimators

def get_estimators_name(estimators):
    result = [ '"{}"'.format(estimator.__class__.__name__) for estimator, _, _ in estimators ]
    return ','.join(result), {estimator_name.strip('"'): estimator for estimator_name, (estimator, _, _) in zip(result, estimators)}

if __name__ == "__main__":
    estimators = get_best_estimators(True)
    estimators_str, estimator_dict = get_estimators_name(estimators)
    print(estimators_str)
    with open('predict.json') as config_file:
        data = json.load(config_file)
        model =     data["model"].format(estimators_str)
        emotions =  data['emotions'].split(",")
        frequency = data["frequency"]
        features =  data["features"].split(",")
        model_ver = data["model_ver"]
        output  =   data["output"]

    

    # Parse the arguments passed
    model_name = os.path.join(model_ver,model)


    # if classifier is SVC need to parse probability as true to display probability
    if(model == "SVC"):
        detector = EmotionRecognizer(model = SVC(probability = True) , emotions=emotions, model_name = model_name,  features=features , verbose=0)

    # similar for decision tree classifier 
    elif(model == "DecisionTreeClassifier"):
        detector = EmotionRecognizer(model = DecisionTreeClassifier() , emotions=emotions, model_name = model_name, features=features, verbose=0)
    
    elif(model == "RNN"):
        detector = DeepEmotionRecognizer(emotions, emodb = True, customdb = True, n_rnn_layers=2, n_dense_layers=2, rnn_units=128, dense_units=128)

    else:
        detector = EmotionRecognizer(estimator_dict[model] , emotions=emotions, model_name = model_name, features=features, verbose=0)
    # record emotions to be predicted 
    emotions = emotions=emotions

    # if form of output is excel
    if(output == "excel"):

        # initialise workbook 
        wb = load_workbook('predict_from_audio/prediction.xlsx')
        wb.remove(wb['predictions'])
        wb.create_sheet('predictions')
        sheet = wb['predictions']
        sheet["A1"] = "emotion to predict"
        sheet["B1"] = "result"

        rows = 2
        cols = 1

        # for emotions to be predicted 
        for i in range(len(emotions)):

            # for the filepath containing that emotion
            for filepath in glob.iglob("predict_from_audio" + os.sep + "emotion testing audio 16k" + os.sep + emotions[i] + os.sep + "/*"):
                # record the emotion in the excel sheet
                sheet[get_column_letter(i + 3) + "1"] =  emotions[i]

                # record prediction probability
                predictions = detector.predict_proba(filepath)

                if(emotions[i]==max(predictions, key=predictions.get).lower()):
                    sheet[get_column_letter(cols) + str(rows)] = str(emotions[i])
                    sheet[get_column_letter(cols + 1) + str(rows)] = "correct"
                    cols += 2        
                else:
                    sheet[get_column_letter(cols) + str(rows)] = str(emotions[i])
                    sheet[get_column_letter(cols + 1) + str(rows)] = "incorrect"
                    cols += 2

                for value in (predictions).values():
                    sheet[get_column_letter(cols) + str(rows)] = value
                    cols += 1
                rows += 1
                cols = 1

        wb.save('predict_from_audio/prediction.xlsx')
        print('predictions saved to predict_from_audio/prediction.xlsx')
    
    else:
        # open file to write predictions on 
        with open(file = 'predict_from_audio' + os.sep + 'predictions.txt', mode  = 'w') as file:

            # iterate through all the files
            file.write("results," + str(emotions) + "\n")
            for emotion in emotions:

                # record emotions to predict to write to the putput file later
                to_predict = emotion + (8-len(emotions))*(" ")

                for filepath in glob.iglob("predict_from_audio" + os.sep + "emotion testing audio 44k" + os.sep + emotions + os.sep + "/*"):
                    
                    # write if prediction was correct
                    if(emotions==detector.predict(filepath).lower()):
                        file.write(to_predict + " correct  :" )
                    else:
                        file.write(to_predict + " incorrect:" )
                    
                    # Write probabiltiy distribution
                    for value in (detector.predict_proba(filepath)).values():
                        file.write(str(value) + ",")
                    file.write(str(filepath) + "\n")
                
        print('predictions saved to predict_from_audio/prediction.txt')
    




